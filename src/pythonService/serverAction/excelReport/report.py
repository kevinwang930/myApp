
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
from serverAction.db import getOrder_db
import importlib
from serverAction.utils import checkDestFileWritable

class OrderExcelReportBase:
    templatePath = 'public/template/orderReport.xlsx'
    orderItemHead = ['产品型号','名称','描述','单价','数量','金额']
    orderNoPosition = "F2"
    supplierNamePosition='B2'
    supplierContactPosition='B3'
    supplierPhonePosition='B4'
    orderItemsRef='A6'

    def __init__(self,orderNo) -> None:
        self.orderNo = orderNo

    def getOrderFromDb(self):
        order = getOrder_db(self.orderNo)
        if order:
            self.orderInfo = order.orderInfo
            self.supplier = order.supplier
            self.orderItems = order.orderItems
            return True
        else:
            return False


class OrderExcelReport_win32com(OrderExcelReportBase):
        
    def run(self):

        
        if self.getOrderFromDb():
            return self.writeToExcel()
        else:
            return False,f'python server get order from db failed'

    def writeToExcel(self):

        self.destFile  = Path(f'output/{self.orderNo}.xlsx').resolve()
        self.xlApp = importlib.import_module('.win32com_interface',package="dbexcel").xl

        self.tryCloseDestFile_win32com()
        self.workbook = self.xlApp.Workbooks.Open(self.templatePath)
        self.worksheet = self.workbook.Worksheets('order')
        
        self.writeOrderInfo()
        self.writeSupplier()
        self.writeOrderItems()
        return self.saveWorkbook()
              
    def writeOrderInfo(self):
        orderNoRange = self.worksheet.Range(self.orderNoPosition)
        orderNoRange.NumberFormat = '@'
        orderNoRange.Value = self.orderNo
    

    def writeSupplier(self):
            supplier = self.supplier
            self.worksheet.Range(self.supplierNamePosition).Value = supplier.name
            self.worksheet.Range(self.supplierContactPosition).Value = supplier.contact
            if (cellphone:=supplier.cellphone) and (telephone:=supplier.telephone):
                self.worksheet.Range(self.supplierPhonePosition).Value = cellphone+' / ' + telephone
            elif cellphone:
                self.worksheet.Range(self.supplierPhonePosition).Value = cellphone
            elif telephone:
                self.worksheet.Range(self.supplierPhonePosition).Value = telephone
                

    def writeOrderItems(self):
        
        column = self.worksheet.Range(self.orderItemsRef).Column
        row = self.worksheet.Range(self.orderItemsRef).Row
        self.worksheet.Range(self.worksheet.Cells(row,column),self.worksheet.Cells(row,column+5)).Value = self.orderItemHead
        for item in self.orderItems:
            row = row + 1
            self.worksheet.Cells(row,column).NumberFormat = '@'
            self.worksheet.Range(self.worksheet.Cells(row,column),self.worksheet.Cells(row,column+4)).Value=item
            priceAddress = self.worksheet.Cells(row,column+3).GetAddress(RowAbsolute=False,ColumnAbsolute=False)
            quantityAddress = self.worksheet.Cells(row,column+4).GetAddress(RowAbsolute=False,ColumnAbsolute=False)
            self.worksheet.Cells(row,column+5).Formula = f'={priceAddress}*{quantityAddress}'

    
  
    def saveWorkbook(self):
      
        fileName = str(self.destFile)
        self.xlApp.DisplayAlerts = False
        try:
            self.workbook.SaveAs(fileName)   
        except:
            for workbook in self.xlApp.Workbooks:
                if workbook.FullName == fileName:
                    workbook.Close()
                    self.workbook.SaveAs(fileName)
                
        self.xlApp.DisplayAlerts = True
        self.xlApp.Visible = True
        return True,f'excel保存成功 {fileName}'

    def tryCloseDestFile_win32com(self):
        for workbook in self.xlApp.Workbooks:
                if workbook.FullName == str(self.destFile):
                    workbook.Close()
                    return


class OrderExcelReport_openpyxl(OrderExcelReportBase):

    
        
    def run(self):
        if self.getOrderFromDb():
            return self.writeToExcel()
        else:
            return False,f'python server get order from db failed'

   
    def writeToExcel(self):
        self.destFile  = Path(f'output/{self.orderNo}.xlsx').resolve()
        if not checkDestFileWritable(self.destFile):
            return False,f'excel {self.destFile}已经打开无法更新'
        self.workbook = load_workbook(self.templatePath)
        self.worksheet = self.workbook['order']
    
        self.writeOrderInfo()
        self.writeSupplier()
        self.writeOrderItems()
        return self.saveWorkbook()

    
                

    def writeOrderInfo(self):
        _cell = self.worksheet[self.orderNoPosition]
        _cell.value=self.orderNo
        _cell.number_format = '@'
       
    
    def writeSupplier(self):
        supplier = self.supplier
        self.worksheet[self.supplierNamePosition].value = supplier.name
        self.worksheet[self.supplierContactPosition].value = supplier.contact
        if (cellphone:=supplier.cellphone) and (telephone:=supplier.telephone):
            self.worksheet[self.supplierPhonePosition].value = cellphone+' / ' + telephone
        elif cellphone:
            self.worksheet[self.supplierPhonePosition].value = cellphone
        elif telephone:
            self.worksheet[self.supplierPhonePosition].value = telephone

    def writeOrderItems(self):
    
        column = self.worksheet[self.orderItemsRef].column
        row = self.worksheet[self.orderItemsRef].row
        for col in range(len(self.orderItemHead)):
            self.worksheet.cell(row,column+col).value = self.orderItemHead[col]
        
        for item in self.orderItems:
            row = row + 1
            self.worksheet.cell(row,column).number_format = '@'
            for col in range(len(self.orderItemHead)-1):
                self.worksheet.cell(row,column+col).value = item[col]

            
            priceAddress = self.worksheet.cell(row,column+3).coordinate
            quantityAddress = self.worksheet.cell(row,column+4).coordinate
            self.worksheet.cell(row,column+5).value = f'={priceAddress}*{quantityAddress}'

  
    
    def saveWorkbook(self):
        dest_file = Path(f'output/{self.orderNo}.xlsx').resolve()
        try:
            self.workbook.save(filename = dest_file)
            os.startfile(dest_file)
            return True,f'excel保存成功 {dest_file}'
        except:
            return False,f'excel {dest_file}写入出错'
 

def orderExcelReport_win32com(orderNo):
    action = OrderExcelReport_win32com(orderNo)
    return action.run()

def orderExcelReport_openpyxl(orderNo):
    action = OrderExcelReport_openpyxl(orderNo)
    return action.run()
   